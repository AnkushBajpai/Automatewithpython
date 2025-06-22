"""
Excel-to-CSV Converter
Excel can save a spreadsheet to a CSV file with a few mouse clicks, 
but if you had to convert hundreds of Excel files to CSVs, 
it would take hours of clicking. Using the openpyxl module from Chapter 12, 
write a program that reads all the Excel files in the current working directory and outputs them as CSV files.

A single Excel file might contain multiple sheets; 
you’ll have to create one CSV file per sheet. 
The filenames of the CSV files should be <excel filename>_<sheet title>.csv, 
where <excel filename> is the filename of the Excel file without the file extension (for example, 'spam_data', not 'spam_data.xlsx') 
and <sheet title> is the string from the Worksheet object’s title variable.

"""
import openpyxl, csv, os

def excel_to_csv(folder):
    for excel_file in os.listdir(folder):

            # Skip non-xlsx files, load the workbook object.
        if not excel_file.endswith('.xlsx'):
            continue
        excel_file_path = os.path.join(folder, excel_file)
        workbook = openpyxl.load_workbook(excel_file_path)
        filename_only = os.path.splitext(os.path.basename(excel_file_path))[0]

                # Loop through every sheet in the workbook.
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

                    # Create the CSV filename from the Excel filename and sheet title.
            csv_file_path = os.path.join(folder, filename_only+'_'+sheet_name + '.csv')
            with open(csv_file_path, 'w', newline='') as csv_file:

                        # Create the csv.writer object for this CSV file.
                csv_writer = csv.writer(csv_file)

                        # Loop through every row in the sheet.
                for row_num in range(1, sheet.max_row + 1):
                    row_values = []   # append each cell to this list

                                # Loop through each cell in the row.
                    for col_num in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row_num, column=col_num).value

                                        # Append each cell's data to rowData.
                        row_values.append(cell_value)

                                    # Write the rowData list to the CSV file.
                    csv_writer.writerow(row_values)

excel_to_csv('E:\python\Automatewithpython\excelToCSV')
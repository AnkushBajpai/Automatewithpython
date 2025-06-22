"""
Write a program to invert the row and column of the cells in the spreadsheet. 
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa). 
This should be done for all cells in the spreadsheet.
"""
from openpyxl import load_workbook

# Open the original workbook and sheet
filename = 'example_copy.xlsx'
wb = load_workbook(filename)
ws = wb.active

# Get the current size of the data
max_row = ws.max_row
max_col = ws.max_column

# Step 1: Read data into sheetData[x][y]
sheetData = [[None for _ in range(max_row)] for _ in range(max_col)]
for y in range(max_row):
    for x in range(max_col):
        sheetData[x][y] = ws.cell(row=y+1, column=x+1).value

# Optional: Clear existing cells (to avoid leftover data if shape shrinks)
for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        cell.value = None

# Step 2: Write transposed data back to same sheet
for x in range(max_col):
    for y in range(max_row):
        ws.cell(row=x+1, column=y+1, value=sheetData[x][y])

# Step 3: Save the file (same file, overwriting it)
wb.save(filename)

print(f"Spreadsheet '{filename}' has been transposed in-place.")

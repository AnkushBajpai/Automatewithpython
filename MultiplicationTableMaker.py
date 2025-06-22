"""
Multiplication Table Maker
Create a program multiplicationTable.py that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet. For example, when the program is run like this:

py multiplicationTable.py 6

Figure 13-11: A multiplication table generated in a spreadsheet

Row 1 and column A should be used for labels and should be in bold.

"""

import sys, openpyxl
from openpyxl.styles import Font

# Check if the script was called with the correct argument
if len(sys.argv) != 2:
    print("Usage: python multiplicationTableMaker.py <number>")
    sys.exit(1)

# Get the number from the command-line argument
number = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active

# populating the sheet with the expected data
for rw in range(2,number+2):
    sheet.cell(row=rw, column=1).value = rw-1
    sheet.cell(row=rw, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=rw).value = rw-1
    sheet.cell(row=1, column=rw).font = Font(bold=True)

for rw in range(2,number+2):
    for col in range(2,number+2):
        sheet.cell(row=rw, column=col).value = sheet.cell(row=1, column=col).value * sheet.cell(row=rw, column=1).value

#saving the sheet in a file
wb.save('MultiplicationTableMakerResults.xlsx')